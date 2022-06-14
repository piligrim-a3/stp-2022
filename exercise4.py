import random
from openpyxl import Workbook
from openpyxl.styles import (Border, Side, Alignment, Font)

#создание объекста класса (ексель файл)
wb = Workbook()
ws = wb.active
#настройка ячеек
for A, e in zip("ABCDEFGHIJ", [5.5, 9.1, 5.5, 22.7, 9.1, 11.2, 13.7, 13.7, 13.8, 13.9]):
    ws.column_dimensions[A].width = e
for A, e in zip([3, 4, 15, 16, 63, 65], [24, 21, 39, 25, 24, 24]):
    ws.row_dimensions[A].height = e
#настройка границ
b = Side(border_style="thin")
for A in "ABCDEFGHIJ":
    for i in range(15, 57):
        ws[A + str(i)].border = Border(top=b, bottom=b, left=b, right=b)
for A in "DEFGHIJ":
    for i in range(57, 58):
        ws[A + str(i)].border = Border(top=b, bottom=b, left=b, right=b)
#объединение ячеек
f = open("Объединение.txt", "r", encoding='utf-8')
while True:
    s = f.readline()
    if s == '':
        break
    ws.merge_cells(s)
#заполнение документа
f = open("Документ.txt", "r", encoding='utf-8')
while True:
    s = f.readline().split('*')
    if s == ['']:
        break
    ws[s[0]].value = s[1].replace('\n', '')
#Выравнивание данных в ячейках
for A in "ABCDEFGHIJ":
    for i in range(1, 67):
        ws[A + str(i)].font = Font(name='Times new roman', size=10)
        ws[A + str(i)].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
#Выравнивание данных в ячейках c именами
for i in range(17, 57):
    ws['D' + str(i)].alignment = Alignment(horizontal="left", vertical="center")
#форматирвоание ячеек с итоговыми суммами
ws['D57'].font, ws['D57'].alignment = Font(name='Times new roman', size=10, bold=True), Alignment(horizontal="right", vertical="center")
for A in "EFGHIJ":
    ws[A + '57'].font, ws[A + '57'].alignment = Font(name='Times new roman', size=10, bold=True), Alignment(horizontal="center", vertical="center")
    #заполнение документа данными
f = open("Имена.txt", "r", encoding='utf-8')