from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Border, Side,
    Alignment, Font, GradientFill,
    Color, colors)

ABC = "ABCDEFGHIJ"
cells = ["I2", "I3", "A10", "A11", "A12", "A13",
         "E63", "E64", "E65", "E66", "H63", "H64", "H65",
         "H66", "A15", "B15", "C15", "D15", "E15", "F16",
         "E16", "G15", "H15", "I15", "J15"]

def data():
    ws['H1'].value = 'УТВЕРЖДАЮ:'
    ws['H2'].value = 'Директор'
    ws['I3'].value = '(сокращенное наименование образовательного учреждения)'
    ws['H4'].value = '_____________'
    ws['H5'].value = '(подпись)'
    ws['I4'].value = '___________________________'
    ws['I5'].value = '(расшифровка подписи)'
    ws['H7'].value = '14.05.2022'
    ws['H8'].value = 'М.П.'
    ws['A10'].value = 'Отчёт о фактическом предоставленном бесплатном питании'
    ws['A11'].value = 'за период с 01.05.2022 по 31.05.2022'
    ws['A13'].value = '(сокращенное наименование образовательного учреждения)'
    ws['A15'].value = '№ п/п'
    ws['B15'].value = '№ счета'
    ws['C15'].value = 'Класс'
    ws['D15'].value = 'Ф.И. ребенка'
    ws['E15'].value = 'Дни посещения'
    ws['E16'].value = 'плановые'
    ws['F16'].value = 'фактические'
    ws['G15'].value = 'Остаток на начало месяца, руб.'
    ws['H15'].value = 'Поступило в текущем месяце на питание, руб.'
    ws['I15'].value = 'Израсходовано в текущем месяце на питание, руб.'
    ws['J15'].value = 'Остаток на конец месяца, руб.'
    ws['D57'].value = 'Итого:'
    ws['B59'].value = 'Отчет составлен в двух экземплярах.'
    ws['B61'].value = 'Подписи сторон:'
    ws['B63'].value = 'Лицо, ответственное за организацию питания'
    ws['B65'].value = 'Заведующий производством предприятия общественного питания'
    ws['E63'].value = '_____________'
    ws['E64'].value = '(подпись)'
    ws['E65'].value = '_____________'
    ws['E66'].value = '(подпись)'
    ws['H63'].value = '___________________________'
    ws['H64'].value = '(Ф.И.О.)'
    ws['H65'].value = '___________________________'
    ws['H66'].value = '(Ф.И.О.)'


def sample():
    ws.merge_cells('A10:J10')
    ws.merge_cells('A11:J11')
    ws.merge_cells('A12:J12')
    ws.merge_cells('A13:J13')
    ws.merge_cells('I2:J2')
    ws.merge_cells('I3:J3')
    ws.merge_cells('I4:J4')
    ws.merge_cells('I5:J5')
    ws.merge_cells('A15:A16')
    ws.merge_cells('B15:B16')
    ws.merge_cells('C15:C16')
    ws.merge_cells('D15:D16')
    ws.merge_cells('E15:F15')
    ws.merge_cells('G15:G16')
    ws.merge_cells('H15:H16')
    ws.merge_cells('I15:I16')
    ws.merge_cells('J15:J16')
    ws.merge_cells('B59:D59')
    ws.merge_cells('B63:D63')
    ws.merge_cells('B65:D65')
    ws.merge_cells('E63:F63')
    ws.merge_cells('E64:F64')
    ws.merge_cells('E65:F65')
    ws.merge_cells('E66:F66')
    ws.merge_cells('H63:I63')
    ws.merge_cells('H64:I64')
    ws.merge_cells('H65:I65')
    ws.merge_cells('H66:I66')


def borders():
    borders = Side(border_style="thin", color="000000")
    for A in ABC:
        for i in range(15, 57):
            ws[A + str(i)].border = Border(top=borders, bottom=borders, left=borders, right=borders)
    for A in "DEFGHIJ":
        for i in range(57, 58):
            ws[A + str(i)].border = Border(top=borders, bottom=borders, left=borders, right=borders)
    ws["I2"].border = Border(bottom=Side(border_style="hair", color="000000"))
    ws["A12"].border = Border(bottom=Side(border_style="thin", color="000000"))


def size():
    w = [5, 9, 6, 22, 9, 11, 14, 14, 14, 14]
    for i in range(10):
        ws.column_dimensions[ABC[i]].width = w[i]
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 21
    ws.row_dimensions[15].height = 39
    ws.row_dimensions[16].height = 25
    ws.row_dimensions[63].height = 24
    ws.row_dimensions[65].height = 24


def align():
    for A in ABC:
        for i in range(1, 67):
            ws[A + str(i)].font = ft
            if A + str(i) in cells:
                ws[A + str(i)].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            else:
                ws[A + str(i)].alignment = Alignment(wrap_text=True)


def format():
    ws['D57'].font = Font(name='Times new roman',
          size=10,
          bold=True,
          italic=False,
          vertAlign=None,
          underline='none',
          strike=False,
          color='FF000000')
    ws['D57'].alignment = Alignment(horizontal="right", vertical="center")


wb = Workbook()
ws = wb.active
ft = Font(name='Times new roman',
          size=10,
          bold=False,
          italic=False,
          vertAlign=None,
          underline='none',
          strike=False,
          color='FF000000')
align()
size()
data()
borders()
sample()
format()
wb.save('test.xlsx')